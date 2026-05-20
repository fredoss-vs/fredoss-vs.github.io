#!/usr/bin/env python3
"""
Build source/taxonomy_db.json depuis CH-ISCO-19 (xlsx) + ESCO v1.2.1 (CSV fr).

Usage:
    python3 build_taxonomy_db.py [--esco-dir <chemin>] [--xlsx <chemin>] [--out <chemin>]

Sources attendues (chemins par défaut relatifs à la racine du skill) :
    exemple_job_search/source/do-f-00.07-ch-isco-19-04.xlsx
    ESCO_dataset _v1.2.1_classification _fr _csv/

Dépendance : openpyxl (pip install openpyxl)

Durée : ~30-60 secondes selon le système.
Taille output : ~1.5 MB (600 groupes ISCO, ~23K métiers CH, skills essentiels).
"""
from __future__ import annotations
import argparse, csv, json, unicodedata, sys, datetime, io
from collections import defaultdict
from pathlib import Path

SKILL_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = SKILL_ROOT / "exemple_job_search" / "source" / "do-f-00.07-ch-isco-19-04.xlsx"
DEFAULT_ESCO = SKILL_ROOT / "ESCO_dataset _v1.2.1_classification _fr _csv"
DEFAULT_OUT  = SKILL_ROOT / "source" / "taxonomy_db.json"

# Skills dont le reuseLevel est retenu pour le matching emploi
KEPT_REUSE = {"cross-sector", "sector-specific"}
MAX_SKILLS_PER_GROUP = 15


def log(msg: str) -> None:
    print(msg, file=sys.stderr, flush=True)


def norm(text: str) -> str:
    if not text:
        return ""
    t = str(text).lower().strip()
    t = unicodedata.normalize("NFD", t)
    return "".join(c for c in t if unicodedata.category(c) != "Mn")


def split_alts(raw: str) -> list[str]:
    """Divise les altLabels ESCO (séparés par \\n) en liste filtrée."""
    if not raw:
        return []
    parts = [p.strip() for p in raw.split("\n")]
    return [p for p in parts if p and len(p) > 2]


def unique_ordered(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out = []
    for item in items:
        k = norm(item)
        if k and k not in seen:
            seen.add(k)
            out.append(item)
    return out


# ──────────────────────────────────────────────────────────────────────────────
# ÉTAPE 1 — CH-ISCO-19 : grouper les métiers par code ISCO à 4 chiffres
# ──────────────────────────────────────────────────────────────────────────────

def load_ch_isco(xlsx_path: Path) -> dict[str, dict]:
    """
    Retourne un dict :
        isco4 → {
            "unit_label": str,
            "major": str, "major_label": str,
            "minor": str, "minor_label": str,
            "ch_names": [str, ...]   # tous les métiers actifs M+F
        }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        log("❌ openpyxl manquant. Installer : pip install openpyxl --break-system-packages")
        sys.exit(1)

    log(f"[1/4] Lecture CH-ISCO-19 : {xlsx_path.name}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["HCL_CH_ISCO_19_PROF_1_2_2"]
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter)

    # Index des colonnes clés
    h = {str(c): i for i, c in enumerate(headers)}
    COL_MAJOR     = h.get("CH_ISCO_19_MAJORGROUP", 0)
    COL_MAJ_LBL   = h.get("Grands_groupes_name", 1)
    COL_MINOR     = h.get("CH_ISCO_19_MINORGROUP", 4)
    COL_MIN_LBL   = h.get("Sous-groupe_name", 5)
    COL_UNIT      = h.get("CH_ISCO_19_UNITGROUP", 6)
    COL_UNIT_LBL  = h.get("Groupes_de_base_name", 7)
    COL_ACTIVE_T  = h.get("ACTIVE_Type", 12)
    COL_ACTIVE_V  = h.get("ACTIVE_Title", 13)
    COL_NAME_FR   = h.get("Name_fr", 11)
    COL_GENDER_F  = h.get("GENDER_F_Text_fr", 21)
    COL_GENDER_M  = h.get("GENDER_M_Text_fr", 23)

    by_isco4: dict[str, dict] = {}
    skipped_inactive = 0

    for row in rows_iter:
        if not row or not any(row):
            continue

        def get(col: int) -> str:
            v = row[col] if col < len(row) else None
            return str(v).strip() if v is not None else ""

        active_type  = get(COL_ACTIVE_T)
        active_value = get(COL_ACTIVE_V)

        # Ne conserver que les entrées actives (ACTIVE=1)
        if active_type != "ACTIVE" or active_value != "1":
            skipped_inactive += 1
            continue

        isco4     = get(COL_UNIT)[:4] if get(COL_UNIT) else ""
        unit_lbl  = get(COL_UNIT_LBL)
        major     = get(COL_MAJOR)
        major_lbl = get(COL_MAJ_LBL)
        minor     = get(COL_MINOR)
        minor_lbl = get(COL_MIN_LBL)
        name_fr   = get(COL_NAME_FR)
        gender_f  = get(COL_GENDER_F)
        gender_m  = get(COL_GENDER_M)

        if not isco4:
            continue

        if isco4 not in by_isco4:
            by_isco4[isco4] = {
                "unit_label": unit_lbl,
                "major": major,
                "major_label": major_lbl,
                "minor": minor,
                "minor_label": minor_lbl,
                "ch_names": [],
            }

        # Collecter les formes de noms
        for raw in [name_fr, gender_f, gender_m]:
            if not raw:
                continue
            # Name_fr peut contenir M | F séparés par " | "
            for part in raw.split("|"):
                part = part.strip()
                if part:
                    by_isco4[isco4]["ch_names"].append(part)

    log(f"   {len(by_isco4)} groupes ISCO-4 | {skipped_inactive} entrées inactives ignorées")
    return by_isco4


# ──────────────────────────────────────────────────────────────────────────────
# ÉTAPE 2 — ESCO skills : charger le dict URI → label + altLabels (FR)
# ──────────────────────────────────────────────────────────────────────────────

def load_esco_skills(esco_dir: Path) -> dict[str, dict]:
    """
    Retourne dict : skillUri → {"label": str, "alts": [str,...], "reuse": str}
    Ne conserve que cross-sector et sector-specific.
    """
    skills_file = esco_dir / "skills_fr.csv"
    log(f"[2/4] Lecture skills ESCO : {skills_file.name}")

    skill_dict: dict[str, dict] = {}
    with open(skills_file, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            reuse = row.get("reuseLevel", "")
            if reuse not in KEPT_REUSE:
                continue
            uri   = row.get("conceptUri", "").strip()
            label = row.get("preferredLabel", "").strip()
            alts  = split_alts(row.get("altLabels", ""))
            if uri and label:
                skill_dict[uri] = {"label": label, "alts": alts, "reuse": reuse}

    log(f"   {len(skill_dict)} skills cross-sector + sector-specific chargés")
    return skill_dict


# ──────────────────────────────────────────────────────────────────────────────
# ÉTAPE 3 — ESCO occupations : charger et regrouper par iscoGroup (4 chiffres)
# ──────────────────────────────────────────────────────────────────────────────

def load_esco_occupations(esco_dir: Path) -> dict[str, dict]:
    """
    Retourne dict : isco4 → {
        "preferred": [str,...],    # preferredLabels des occupations de ce groupe
        "alts": [str,...],         # tous les altLabels du groupe
        "uris": [str,...],         # URIs pour le join avec skills
    }
    """
    occ_file = esco_dir / "occupations_fr.csv"
    log(f"[3/4] Lecture occupations ESCO : {occ_file.name}")

    by_isco4: dict[str, dict] = {}
    with open(occ_file, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            isco4 = (row.get("iscoGroup") or "").strip()[:4]
            uri   = (row.get("conceptUri") or "").strip()
            pref  = (row.get("preferredLabel") or "").strip()
            alts  = split_alts(row.get("altLabels", ""))

            if not isco4:
                continue

            if isco4 not in by_isco4:
                by_isco4[isco4] = {"preferred": [], "alts": [], "uris": []}

            if pref:
                by_isco4[isco4]["preferred"].append(pref)
            by_isco4[isco4]["alts"].extend(alts)
            if uri:
                by_isco4[isco4]["uris"].append(uri)

    log(f"   {len(by_isco4)} groupes ISCO-4 ESCO")
    return by_isco4


# ──────────────────────────────────────────────────────────────────────────────
# ÉTAPE 4 — Relations occupation → skills : construire map URI → top skills FR
# ──────────────────────────────────────────────────────────────────────────────

def load_occupation_skills(
    esco_dir: Path,
    skill_dict: dict[str, dict],
    esco_occs: dict[str, dict],
) -> dict[str, list[str]]:
    """
    Retourne dict : isco4 → [skill_label_fr, ...]  (top MAX_SKILLS_PER_GROUP labels)
    En groupant par ISCO-4 via les URIs d'occupation.
    """
    rel_file = esco_dir / "occupationSkillRelations_fr.csv"
    log(f"[4/4] Lecture relations occupation-skill : {rel_file.name}")

    # Construire reverse map : occUri → isco4
    uri_to_isco4: dict[str, str] = {}
    for isco4, data in esco_occs.items():
        for uri in data["uris"]:
            uri_to_isco4[uri] = isco4

    # Compter les occurrences de chaque skill par groupe ISCO-4
    skill_counts: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    processed = 0

    with open(rel_file, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rel_type = row.get("relationType", "")
            if rel_type != "essential":
                continue
            occ_uri   = (row.get("occupationUri") or "").strip()
            skill_uri = (row.get("skillUri") or "").strip()

            isco4 = uri_to_isco4.get(occ_uri)
            if not isco4:
                continue

            skill = skill_dict.get(skill_uri)
            if not skill:
                continue

            skill_label = skill["label"]
            skill_counts[isco4][skill_label] += 1
            processed += 1

    log(f"   {processed} relations essentielles traitées")

    # Garder les top N skills par fréquence dans le groupe
    isco4_to_skills: dict[str, list[str]] = {}
    for isco4, counts in skill_counts.items():
        top = sorted(counts.items(), key=lambda x: x[1], reverse=True)
        isco4_to_skills[isco4] = [label for label, _ in top[:MAX_SKILLS_PER_GROUP]]

    return isco4_to_skills


# ──────────────────────────────────────────────────────────────────────────────
# ASSEMBLAGE FINAL
# ──────────────────────────────────────────────────────────────────────────────

def build_db(
    ch_data: dict[str, dict],
    esco_occs: dict[str, dict],
    isco4_skills: dict[str, list[str]],
) -> dict:
    """Assemble le dict final taxonomy_db."""
    all_isco4 = set(ch_data.keys()) | set(esco_occs.keys())
    db: dict[str, dict] = {}

    for isco4 in sorted(all_isco4):
        ch   = ch_data.get(isco4, {})
        esco = esco_occs.get(isco4, {})

        ch_names  = unique_ordered(ch.get("ch_names", []))
        esco_pref = unique_ordered(esco.get("preferred", []))
        esco_alts = unique_ordered(esco.get("alts", []))
        skills    = isco4_skills.get(isco4, [])

        # Exclure les altLabels déjà présents dans ch_names ou esco_pref
        existing_norms = {norm(x) for x in ch_names + esco_pref}
        esco_alts = [a for a in esco_alts if norm(a) not in existing_norms]

        db[isco4] = {
            "isco4":       isco4,
            "unit_label":  ch.get("unit_label") or (esco_pref[0] if esco_pref else ""),
            "major":       ch.get("major", ""),
            "major_label": ch.get("major_label", ""),
            "minor":       ch.get("minor", ""),
            "minor_label": ch.get("minor_label", ""),
            "ch_names":    ch_names,
            "esco_preferred": esco_pref,
            "esco_alts":   esco_alts,
            "skills":      skills,
        }

    return db


def main() -> None:
    parser = argparse.ArgumentParser(description="Build swiss-job-intel taxonomy DB")
    parser.add_argument("--xlsx",     type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--esco-dir", type=Path, default=DEFAULT_ESCO)
    parser.add_argument("--out",      type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    if not args.xlsx.exists():
        log(f"❌ CH-ISCO xlsx non trouvé : {args.xlsx}")
        sys.exit(1)
    if not args.esco_dir.exists():
        log(f"❌ Répertoire ESCO non trouvé : {args.esco_dir}")
        sys.exit(1)

    log("🔨 Construction de taxonomy_db.json")
    log(f"   CH-ISCO  : {args.xlsx}")
    log(f"   ESCO CSV : {args.esco_dir}")
    log(f"   Output   : {args.out}")
    log("")

    ch_data    = load_ch_isco(args.xlsx)
    skill_dict = load_esco_skills(args.esco_dir)
    esco_occs  = load_esco_occupations(args.esco_dir)
    isco4_sk   = load_occupation_skills(args.esco_dir, skill_dict, esco_occs)

    log("")
    log("Assemblage final...")
    db = build_db(ch_data, esco_occs, isco4_sk)

    # Statistiques
    total_ch    = sum(len(v["ch_names"]) for v in db.values())
    total_alts  = sum(len(v["esco_alts"]) for v in db.values())
    total_skills = sum(len(v["skills"]) for v in db.values())

    output = {
        "_meta": {
            "built":           datetime.date.today().isoformat(),
            "source_ch_isco":  str(args.xlsx.name),
            "source_esco":     "ESCO v1.2.1 fr",
            "isco4_groups":    len(db),
            "total_ch_names":  total_ch,
            "total_esco_alts": total_alts,
            "total_skills":    total_skills,
        },
        "by_isco4": db,
    }

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = args.out.stat().st_size // 1024
    log(f"\n✅ taxonomy_db.json écrit ({size_kb} KB)")
    log(f"   {len(db)} groupes ISCO-4")
    log(f"   {total_ch} noms de métiers CH-ISCO")
    log(f"   {total_alts} synonymes ESCO")
    log(f"   {total_skills} compétences essentielles")
    log(f"   → {args.out}")


if __name__ == "__main__":
    main()
